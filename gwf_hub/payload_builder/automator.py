from openpyxl import Workbook
from openpyxl import load_workbook
xml_string = ''
dict_string = 'var workflow_structure = [\n['


def readingXMLFile():
    global xml_string
    line = ''
    # Formatting the XML Code for better code alignment
    with open(file='bpmn-digital.txt', encoding='UTF-8', mode='r') as xml_file:
        for lines in xml_file:
            for words in lines:
                if words != '>':
                    line += words
                if words == '>':
                    line = line + '>'
                    xml_string += line + '\n'
                    line = ''
        xml_file.close()

    with open(file='bpmn-digital.txt', encoding='UTF-8', mode='w') as f:
        f.write(xml_string)
        f.close()
    # XML code formatted and saved in the same .txt file

    extractingData()


def extractingData():
    global xml_string, dict_string
    wb = load_workbook("Answer_Automation.xlsx")
    ws = wb["Sheet"]
    step = ''
    stop_clock = 0
    row_counter = 2
    question = ''
    question_id = ''
    step = ''
    found_initial = False
    initial_section_found = 0
    prev_section = ''
    ws.cell(1, 1).value = "Workflow Level"
    ws.cell(1, 2).value = "Workflow Name"
    ws.cell(1, 3).value = "Step"
    ws.cell(1, 4).value = "Question String"
    ws.cell(1, 5).value = "Answer ID"
    with open(file='bpmn-digital.txt', encoding='UTF-8', mode='r') as xml_file:
        for lines in xml_file:
            lines = str(lines)
            if '<bpmn:userTask id=' in lines:
                found_initial = True
                # Capturing the task name
                lines = lines.split('"')
                if '&#10;_VIP' in lines[3]:
                    step = lines[3].rstrip('&#10;_VIP\\')
                elif '&#10;' in lines[3]:
                    step = lines[3].replace('&#10;', '')
                    step = step.replace('\\', '')
                else:
                    step = lines[3].rstrip('\\')

                if initial_section_found == 1:
                    if prev_section != step and step != '':
                        dict_string += '\n],\n['
                        prev_section = step
                    prev_section = step
                if initial_section_found == 0:
                    prev_section = step
                    initial_section_found = 1

                ws.cell(row_counter, 3).value = step
                wb.save("Answer_Automation.xlsx")
            if 'SELECT_BUTTON' in lines or 'SELECT_RADIO' in lines:
                lines = lines.split('"')
                question_id = lines[1].rstrip('\\')
                question = lines[3].rstrip('\\')
                ws.cell(row_counter, 4).value = question
                ws.cell(row_counter, 5).value = question_id
                ws.cell(row_counter, 3).value = step
                wb.save("Answer_Automation.xlsx")
                row_counter += 1
                type_ans = "string"
                js_dict(prev_section, step, question, question_id, type_ans)
            if 'YES_NO_QUESTION' in lines:
                lines = lines.split('"')
                question_id = lines[1].rstrip('\\')
                question = lines[3].rstrip('\\')
                ws.cell(row_counter, 4).value = question
                ws.cell(row_counter, 5).value = question_id
                ws.cell(row_counter, 3).value = step
                wb.save("Answer_Automation.xlsx")
                row_counter += 1
                type_ans = "bool"
                js_dict(prev_section, step, question, question_id, type_ans)
            if 'SELECT_ONE' in lines:
                lines = lines.split('"')
                question_id = lines[1].rstrip('\\')
                question = lines[3].rstrip('\\')
                ws.cell(row_counter, 4).value = question
                ws.cell(row_counter, 5).value = question_id
                ws.cell(row_counter, 3).value = step
                wb.save("Answer_Automation.xlsx")
                row_counter += 1
                type_ans = "string"
                js_dict(prev_section, step, question, question_id, type_ans)
            if 'CHECKBOX' in lines:
                lines = lines.split('"')
                question_id = lines[1].rstrip('\\')
                question = lines[3].rstrip('\\')
                ws.cell(row_counter, 4).value = question
                ws.cell(row_counter, 5).value = question_id
                ws.cell(row_counter, 3).value = step
                wb.save("Answer_Automation.xlsx")
                row_counter += 1
                type_ans = "bool"
                js_dict(prev_section, step, question, question_id, type_ans)

            if lines == 'bpmn:definitions>"':
                row_counter += 2
                found_initial = False
                if found_initial is False:
                    stop_clock += 1
                    wb.save("Answer_Automation.xlsx")
                    # wb.close()
                    if stop_clock > 10:
                        break
    dict_string += '\n]\n]'
    wb.close()
    print(dict_string)
    with open('aa2.js', mode='w') as js_file:
        js_file.write(dict_string)
        js_file.close()


def js_dict(prev_section, step, question, question_id, type_ans):
    global dict_string
    if prev_section == step:
        if question != '':
            dict_string += '\n{ q_label: "' + question + '",\n'  # Adding question string
            dict_string += 'q_id: "' + question_id + '",\n'  # Adding question id
            dict_string += ' step: "' + prev_section + '",\n'  # Adding question string
            dict_string += ' type: "' + type_ans + '"\n},'  # Adding question string


def creatingEmptyExcel():
    wb = Workbook()
    ws = wb.active
    wb.save("Answer_Automation.xlsx")
    wb.close()


creatingEmptyExcel()
readingXMLFile()
