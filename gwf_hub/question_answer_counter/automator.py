from openpyxl import Workbook
from openpyxl import load_workbook
xml_string = ''


def readingXMLFile():
    global xml_string
    line = ''
    # Formatting the XML Code for better code alignment
    with open(file='bpmn-digital.txt', encoding='UTF-8', mode='r') as xml_file:
        for lines in xml_file:
            for words in lines:
                if words != '>':
                    line += words
                if words == '>':
                    line = line + '>'
                    xml_string += line + '\n'
                    line = ''
        xml_file.close()

    with open(file='bpmn-digital.txt', encoding='UTF-8', mode='w') as f:
        f.write(xml_string)
        f.close()
    # XML code formatted and saved in the same .txt file

    extractingData()


def extractingData():
    global xml_string
    wb = load_workbook("Answer_Automation.xlsx")
    ws = wb["Sheet"]
    step = ''
    row_counter = 2
    ws.cell(1, 1).value = "Workflow Level"
    ws.cell(1, 2).value = "Workflow Name"
    ws.cell(1, 3).value = "Step"
    ws.cell(1, 4).value = "Question String"
    ws.cell(1, 5).value = "Answer ID"
    with open(file='bpmn-digital.txt', encoding='UTF-8', mode='r') as xml_file:
        for lines in xml_file:
            lines = str(lines)
            if '<bpmn:userTask id=' in lines:
                # Capturing the task name
                lines = lines.split('"')
                if '&#10;_VIP' in lines[3]:
                    step = lines[3].rstrip('&#10;_VIP\\')
                elif '&#10;' in lines[3]:
                    step = lines[3].rstrip('&#10;\\')
                else:
                    step = lines[3].rstrip('\\')
                print("Section name : " + step)
                ws.cell(row_counter, 3).value = step
                wb.save("Answer_Automation.xlsx")
            if 'YES_NO_QUESTION' in lines or 'SELECT_BUTTON' in lines or 'SELECT_RADIO' in lines:
                lines = lines.split('"')
                question_id = lines[1].rstrip('\\')
                print(question_id)
                question = lines[3].rstrip('\\')
                ws.cell(row_counter, 4).value = question
                ws.cell(row_counter, 5).value = question_id
                ws.cell(row_counter, 3).value = step
                wb.save("Answer_Automation.xlsx")
                row_counter += 1
            if 'SELECT_ONE' in lines or 'CHECKBOX' in lines:
                lines = lines.split('"')
                question_id = lines[1].rstrip('\\')
                print(question_id)
                question = lines[3].rstrip('\\')
                ws.cell(row_counter, 4).value = question
                ws.cell(row_counter, 5).value = question_id
                ws.cell(row_counter, 3).value = step
                wb.save("Answer_Automation.xlsx")
                row_counter += 1
            if lines == 'bpmn:definitions>"':
                wb.save("Answer_Automation.xlsx")
                # wb.close()
                print("Code executed!")
                break
            wb.save("Answer_Automation.xlsx")

    wb.close()


def creatingEmptyExcel():
    wb = Workbook()
    ws = wb.active
    wb.save("Answer_Automation.xlsx")
    wb.close()


creatingEmptyExcel()
readingXMLFile()
